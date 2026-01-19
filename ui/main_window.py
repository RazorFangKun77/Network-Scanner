from PySide6.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                               QLabel, QLineEdit, QPushButton, QTableWidget,
                               QTableWidgetItem, QProgressBar, QTabWidget, QSpinBox, 
                               QComboBox, QDateEdit, QFileDialog, QMessageBox, QMenu)
from PySide6.QtCore import QThreadPool, QTimer, QDate, Qt
from PySide6.QtGui import QPalette, QColor, QKeySequence, QShortcut
from PySide6.QtWidgets import QApplication
from controllers.scanner import ScanWorker
from models.database import Database
from datetime import datetime
import csv
import sys
import re


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Network Scanner & Ping Monitor")
        self.setGeometry(100, 100, 1000, 700)
        
        self.thread_pool = QThreadPool()
        #Dibatasin 50 wok biar ngga overload,nanti crash njir
        self.thread_pool.setMaxThreadCount(50)  #Maksimal 50 thread ajah(jangan banyak2)
        
        self.db = Database()
        self.scanning = False
        self.monitoring = False
        self.monitor_ips = []
        self.dark_mode = False
        self.scan_session_id = 0
        self.scanned_count = 0
        self.online_count = 0
        self.current_scan_total = 0
        self._pending_log_refresh = False  #Biar ngga sering refresh log
        
        self.init_ui()
        self.setup_shortcuts()
        self.load_logs()
    
    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(10, 5, 10, 10)  #Buat Margin
        main_layout.setSpacing(5)  #Buat Spacing
        central_widget.setLayout(main_layout)
        
        #Buat Toggle Dark Mode
        header_layout = QHBoxLayout()
        header_layout.setSpacing(10)
        
        title_label = QLabel("Network Scanner & Ping Monitor")
        title_label.setStyleSheet("font-size: 16px; font-weight: bold;")
        header_layout.addWidget(title_label)
        
        header_layout.addStretch()
        
        self.dark_mode_button = QPushButton("🌙 Dark Mode")
        self.dark_mode_button.clicked.connect(self.toggle_dark_mode)
        self.dark_mode_button.setFixedWidth(130)
        self.dark_mode_button.setFixedHeight(30)
        header_layout.addWidget(self.dark_mode_button)
        
        main_layout.addLayout(header_layout)
        
        #Tab Widget
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)
        
        #Scanner Tab
        scanner_tab = QWidget()
        scanner_layout = QVBoxLayout()
        scanner_layout.setContentsMargins(10, 10, 10, 10)
        scanner_layout.setSpacing(8)
        scanner_tab.setLayout(scanner_layout)
        
        #Input section
        input_layout = QHBoxLayout()
        input_layout.setSpacing(8)
        input_layout.addWidget(QLabel("IP Range:"))
        
        self.ip_input = QLineEdit()
        self.ip_input.setPlaceholderText("192.168.1.x-xxx")
        input_layout.addWidget(self.ip_input)
        
        self.scan_button = QPushButton("Start Scan")
        self.scan_button.clicked.connect(self.toggle_scan)
        self.scan_button.setFixedWidth(120)
        input_layout.addWidget(self.scan_button)
        
        scanner_layout.addLayout(input_layout)
        
        #Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setFixedHeight(20)
        scanner_layout.addWidget(self.progress_bar)
        
        #Tabel Hasil Scan
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["IP Address", "Hostname", "MAC Address", "OS Type", "Status", "Latency"])
        self.table.horizontalHeader().setStretchLastSection(True)
        #Biar Semua Element di Tengah secara default
        self.table.setProperty("alignment", "center")
        #Biar bisa Pilih Banyak Sekaligus
        self.table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        #Menu Konteks(klik Kanan)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_scanner_context_menu)
        scanner_layout.addWidget(self.table)
        
        # Status label(di bawah)
        self.status_label = QLabel("Ready")
        scanner_layout.addWidget(self.status_label)
        
        self.tab_widget.addTab(scanner_tab, "Scanner")
        
        #Tab Ping Monitor
        monitor_tab = QWidget()
        monitor_layout = QVBoxLayout()
        monitor_layout.setContentsMargins(10, 10, 10, 10)
        monitor_layout.setSpacing(8)
        monitor_tab.setLayout(monitor_layout)
        
        #Tab Monitor IP
        monitor_control = QHBoxLayout()
        monitor_control.setSpacing(8)
        monitor_control.addWidget(QLabel("Monitor IP:"))
        
        self.monitor_ip_input = QLineEdit()
        self.monitor_ip_input.setPlaceholderText("192.168.1.x")
        self.monitor_ip_input.setMaximumWidth(150)
        monitor_control.addWidget(self.monitor_ip_input)
        
        monitor_control.addWidget(QLabel("Interval (sec):"))
        self.interval_input = QSpinBox()
        self.interval_input.setMinimum(1)
        self.interval_input.setMaximum(300)
        self.interval_input.setValue(5)
        self.interval_input.setFixedWidth(70)
        monitor_control.addWidget(self.interval_input)
        
        self.add_monitor_button = QPushButton("Add to Monitor")
        self.add_monitor_button.clicked.connect(self.add_to_monitor)
        self.add_monitor_button.setFixedWidth(120)
        monitor_control.addWidget(self.add_monitor_button)
        
        self.start_monitor_button = QPushButton("Start Monitoring")
        self.start_monitor_button.clicked.connect(self.toggle_monitoring)
        self.start_monitor_button.setFixedWidth(130)
        monitor_control.addWidget(self.start_monitor_button)
        
        self.remove_monitor_button = QPushButton("Remove from Monitor")
        self.remove_monitor_button.clicked.connect(self.remove_from_monitor)
        self.remove_monitor_button.setFixedWidth(150)
        monitor_control.addWidget(self.remove_monitor_button)
        
        monitor_control.addStretch()
        monitor_layout.addLayout(monitor_control)
        
        #Tabel Monitor IP
        self.monitor_table = QTableWidget()
        self.monitor_table.setColumnCount(6)
        self.monitor_table.setHorizontalHeaderLabels(["IP Address", "Hostname", "MAC Address", "Status", "Latency", "Last Check"])
        self.monitor_table.horizontalHeader().setStretchLastSection(True)
        #Biar bisa Pilih Banyak Sekaligus
        self.monitor_table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.monitor_table.setSelectionBehavior(QTableWidget.SelectRows)
        #Menu Konteks(klik Kanan)
        self.monitor_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.monitor_table.customContextMenuRequested.connect(self.show_monitor_context_menu)
        monitor_layout.addWidget(self.monitor_table)
        
        #Status Label
        self.monitor_status_label = QLabel("Not monitoring")
        monitor_layout.addWidget(self.monitor_status_label)
        
        self.tab_widget.addTab(monitor_tab, "Ping Monitor")
        
        #Tab Log Viewer(Hasil Scan dan Monitor)
        log_tab = QWidget()
        log_layout = QVBoxLayout()
        log_layout.setContentsMargins(10, 10, 10, 10)
        log_layout.setSpacing(8)
        log_tab.setLayout(log_layout)
        
        #Buat Ngefilter Log
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(8)
        
        filter_layout.addWidget(QLabel("Filter IP:"))
        self.filter_ip_input = QLineEdit()
        self.filter_ip_input.setPlaceholderText("e.g., 192.168.1")
        self.filter_ip_input.setMaximumWidth(150)
        self.filter_ip_input.textChanged.connect(self.apply_filters)
        filter_layout.addWidget(self.filter_ip_input)
        
        filter_layout.addWidget(QLabel("Status:"))
        self.filter_status_combo = QComboBox()
        self.filter_status_combo.addItems(["All", "Online", "Offline"])
        self.filter_status_combo.setFixedWidth(100)
        self.filter_status_combo.currentTextChanged.connect(self.apply_filters)
        filter_layout.addWidget(self.filter_status_combo)
        
        filter_layout.addWidget(QLabel("Date:"))
        self.filter_date_input = QDateEdit()
        self.filter_date_input.setCalendarPopup(True)
        self.filter_date_input.setDate(QDate.currentDate())
        self.filter_date_input.setSpecialValueText("All Dates")
        self.filter_date_input.setFixedWidth(120)
        self.filter_date_input.dateChanged.connect(self.apply_filters)
        filter_layout.addWidget(self.filter_date_input)
        
        self.clear_filter_button = QPushButton("Clear Filters")
        self.clear_filter_button.clicked.connect(self.clear_filters)
        self.clear_filter_button.setFixedWidth(100)
        filter_layout.addWidget(self.clear_filter_button)
        
        filter_layout.addStretch()
        log_layout.addLayout(filter_layout)
        
        # Tombol Buat Export
        export_layout = QHBoxLayout()
        export_layout.setSpacing(8)
        export_layout.addStretch()
        
        self.clear_history_button = QPushButton("🗑️ Clear History")
        self.clear_history_button.clicked.connect(self.clear_log_history)
        self.clear_history_button.setFixedWidth(140)
        self.clear_history_button.setStyleSheet("QPushButton { background-color: #ff4444; color: white; font-weight: bold; }")
        export_layout.addWidget(self.clear_history_button)
        
        self.export_csv_button = QPushButton("📄 Export to CSV")
        self.export_csv_button.clicked.connect(self.export_to_csv)
        self.export_csv_button.setFixedWidth(140)
        export_layout.addWidget(self.export_csv_button)
        
        self.export_excel_button = QPushButton("📊 Export to Excel")
        self.export_excel_button.clicked.connect(self.export_to_excel)
        self.export_excel_button.setFixedWidth(140)
        export_layout.addWidget(self.export_excel_button)
        
        log_layout.addLayout(export_layout)
        
        #Tabel Log Viewer
        self.log_table = QTableWidget()
        self.log_table.setColumnCount(7)
        self.log_table.setHorizontalHeaderLabels(["Timestamp", "IP Address", "Hostname", "MAC Address", "OS Type", "Status", "Latency"])
        self.log_table.horizontalHeader().setStretchLastSection(True)
        #Menu Konteks(klik Kanan)
        self.log_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.log_table.customContextMenuRequested.connect(self.show_log_context_menu)
        log_layout.addWidget(self.log_table)
        
        self.tab_widget.addTab(log_tab, "Log Viewer")
        
        #Biar bisa atur waktu monitoring
        self.monitor_timer = QTimer()
        self.monitor_timer.timeout.connect(self.run_monitor_scan)
    
    def setup_shortcuts(self):
        """Setup keyboard shortcuts"""
        # Ctrl+A untuk Select All di scanner table
        scanner_select_all = QShortcut(QKeySequence("Ctrl+A"), self.table)
        scanner_select_all.activated.connect(self.table.selectAll)
        
        # Ctrl+A untuk Select All di monitor table
        monitor_select_all = QShortcut(QKeySequence("Ctrl+A"), self.monitor_table)
        monitor_select_all.activated.connect(self.monitor_table.selectAll)
        
        # Ctrl+A untuk Select All di log table
        log_select_all = QShortcut(QKeySequence("Ctrl+A"), self.log_table)
        log_select_all.activated.connect(self.log_table.selectAll)
    
    def toggle_dark_mode(self):
        self.dark_mode = not self.dark_mode
        
        if self.dark_mode:
            #Config Dark Mode
            dark_palette = QPalette()
            dark_palette.setColor(QPalette.Window, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.WindowText, QColor(255, 255, 255))
            dark_palette.setColor(QPalette.Base, QColor(35, 35, 35))
            dark_palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.ToolTipBase, QColor(25, 25, 25))
            dark_palette.setColor(QPalette.ToolTipText, QColor(255, 255, 255))
            dark_palette.setColor(QPalette.Text, QColor(255, 255, 255))
            dark_palette.setColor(QPalette.Button, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.ButtonText, QColor(255, 255, 255))
            dark_palette.setColor(QPalette.BrightText, QColor(255, 0, 0))
            dark_palette.setColor(QPalette.Link, QColor(42, 130, 218))
            dark_palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
            dark_palette.setColor(QPalette.HighlightedText, QColor(255, 255, 255))
            
            self.setPalette(dark_palette)
            self.dark_mode_button.setText("☀️ Light Mode")
            
            #stylesheet biar cocok di dark mode
            self.setStyleSheet("""
                QTableWidget {
                    gridline-color: #666666;
                    background-color: #353535;
                    color: #ffffff;
                }
                QHeaderView::section {
                    background-color: #454545;
                    color: #ffffff;
                    padding: 4px;
                    border: 1px solid #666666;
                }
                QLineEdit, QSpinBox, QDateEdit, QComboBox {
                    background-color: #454545;
                    color: #ffffff;
                    border: 1px solid #666666;
                }
            """)
        else:
            #Mode Terang (Default)
            self.setPalette(self.style().standardPalette())
            self.dark_mode_button.setText("🌙 Dark Mode")
            
            #Reset stylesheet dan atur warna mode terang
            self.setStyleSheet("""
                QTableWidget {
                    gridline-color: #d0d0d0;
                    background-color: #ffffff;
                    color: #000000;
                }
                QHeaderView::section {
                    background-color: #f0f0f0;
                    color: #000000;
                    padding: 4px;
                    border: 1px solid #d0d0d0;
                    font-weight: bold;
                }
                QLabel {
                    color: #000000;
                }
                QLineEdit, QSpinBox, QDateEdit, QComboBox {
                    background-color: #ffffff;
                    color: #000000;
                    border: 1px solid #cccccc;
                }
                QPushButton {
                    background-color: #f0f0f0;
                    color: #000000;
                    border: 1px solid #cccccc;
                    padding: 5px;
                }
                QPushButton:hover {
                    background-color: #e0e0e0;
                }
            """)
    
    def apply_filters(self):
        ip_filter = self.filter_ip_input.text().strip()
        status_filter = self.filter_status_combo.currentText()
        
        #Buat ngambil filter tanggal
        date_filter = ""
        if self.filter_date_input.text() != "All Dates":
            date_filter = self.filter_date_input.date().toString("yyyy-MM-dd")
        
        #Ambil log yang sudah difilter
        logs = self.db.fetch_logs_filtered(ip_filter, status_filter, date_filter)
        
        #Update tabel
        self.log_table.setRowCount(0)
        for log in logs:
            row = self.log_table.rowCount()
            self.log_table.insertRow(row)
            for col, value in enumerate(log):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignCenter)
                self.log_table.setItem(row, col, item)
    #Fungsi buat ngeclear filter
    def clear_filters(self):
        self.filter_ip_input.clear()
        self.filter_status_combo.setCurrentIndex(0)
        self.filter_date_input.setDate(QDate.currentDate())
        self.filter_date_input.setSpecialValueText("All Dates")
        self.load_logs()
    #Fungsi Export ke CSV/Excel
    def export_to_csv(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Export to CSV", 
            "network_log.csv", 
            "CSV Files (*.csv)"
        )
        #Kalau user milih Direktori buat nyimpan
        if file_path:
            try:
                with open(file_path, 'w', newline='', encoding='utf-8') as file:
                    writer = csv.writer(file)
                    
                    # Write headers
                    headers = []
                    for col in range(self.log_table.columnCount()):
                        headers.append(self.log_table.horizontalHeaderItem(col).text())
                    writer.writerow(headers)
                    
                    # Write data
                    for row in range(self.log_table.rowCount()):
                        row_data = []
                        for col in range(self.log_table.columnCount()):
                            item = self.log_table.item(row, col)
                            row_data.append(item.text() if item else "")
                        writer.writerow(row_data)
                
                QMessageBox.information(self, "Success", f"Data exported to {file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to export: {str(e)}")
    #Fungsi Export ke Excel
    def export_to_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            #Jangan install otomatis(Takutnya Ngebug)! Minta user install manual ajah
            QMessageBox.critical(
                self, 
                "Module Not Found", 
                "The 'openpyxl' module is required for Excel export.\n\n"
                "Please install it manually using:\n"
                "pip install openpyxl\n\n"
                "Or add it to your requirements.txt file."
            )
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Export to Excel", 
            "network_log.xlsx", 
            "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Network Log"
                
                # Write headers with styling
                headers = []
                for col in range(self.log_table.columnCount()):
                    headers.append(self.log_table.horizontalHeaderItem(col).text())
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                
                # Write data
                for row in range(self.log_table.rowCount()):
                    for col in range(self.log_table.columnCount()):
                        item = self.log_table.item(row, col)
                        ws.cell(row=row+2, column=col+1, value=item.text() if item else "")
                
                # Auto-adjust column width
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(file_path)
                QMessageBox.information(self, "Success", f"Data exported to {file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to export: {str(e)}")
# Fungsi buat ngehandle tombol Start/Stop Scan
    def toggle_scan(self):
        if not self.scanning:
            self.start_scan()
        else:
            self.stop_scan()
    # Fungsi buat mulai scan
    def start_scan(self):
        ip_range = self.ip_input.text().strip()
        if not ip_range:
            QMessageBox.warning(self, "Input Error", "Please enter IP range")
            self.status_label.setText("Error: Please enter IP range")
            return
        
        try:
            # Validasi dan parse IP range(1.3.1-254 atau 1.3.10-20)
            ip_pattern = r'^(\d{1,3})\.(\d{1,3})\.(\d{1,3})\.(\d{1,3})(?:-(\d{1,3}))?$'
            match = re.match(ip_pattern, ip_range)
            
            if not match:
                QMessageBox.critical(
                    self, 
                    "Invalid Format", 
                    "Invalid IP format!\n\n"
                    "Valid formats:\n"
                    "• Single IP: 192.168.1.1\n"
                    "• IP Range: 192.168.1.1-254\n"
                    "• Short Range: 192.168.1.10-20"
                )
                self.status_label.setText("Error: Invalid IP format")
                return
            
            #identifikasi octet dan range IP
            octet1, octet2, octet3, octet4, range_end = match.groups()
            
            #Validasi octet (0-255)
            for octet in [octet1, octet2, octet3, octet4]:
                if int(octet) > 255:
                    QMessageBox.critical(
                        self, 
                        "Invalid IP", 
                        f"IP octet value {octet} is invalid!\n"
                        "Each octet must be between 0-255."
                    )
                    self.status_label.setText("Error: Invalid IP octet value")
                    return
            
            base_ip = f"{octet1}.{octet2}.{octet3}"
            start = int(octet4)
            
            #Buat validasi range end
            if range_end:
                end = int(range_end)
                
                if end > 255:
                    QMessageBox.critical(
                        self, 
                        "Invalid Range", 
                        f"Range end value {end} is invalid!\n"
                        "Must be between 0-255."
                    )
                    self.status_label.setText("Error: Invalid range end value")
                    return
                
                if start > end:
                    QMessageBox.critical(
                        self, 
                        "Invalid Range", 
                        f"Start value ({start}) cannot be greater than end value ({end})!"
                    )
                    self.status_label.setText("Error: Invalid range")
                    return
            else:
                end = start
            
            total = end - start + 1
            if total > 254:
                reply = QMessageBox.question(
                    self,
                    "Large Scan Warning",#Peringatan Scan Besar
                    f"You are about to scan {total} IP addresses.\n"
                    "This may take a long time.\n\n"
                    "Do you want to continue?",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
                if reply == QMessageBox.No:#Kalau user pilih No
                    self.status_label.setText("Scan cancelled")
                    return
            
            # Increment(penambahan) scan session ID
            self.scan_session_id += 1
            
            self.scanning = True
            self.scan_button.setText("Stop Scan")
            self.table.setRowCount(0)
            
            #Buat reset progress bar dan counter
            self.current_scan_total = total
            self.progress_bar.setMaximum(total)
            self.progress_bar.setValue(0)
            self.scanned_count = 0
            self.online_count = 0
            
            self.status_label.setText(f"Scanning {total} IP addresses...")
            
            # Start scanning - pass session_id ke worker, bukan pakai lambda
            for i in range(start, end + 1):
                ip = f"{base_ip}.{i}"
                worker = ScanWorker(ip, self.scan_session_id)
                worker.signals.result.connect(self.on_scan_result)
                self.thread_pool.start(worker)
        
        except ValueError as e:
            QMessageBox.critical(
                self, 
                "Parsing Error", 
                f"Failed to parse IP range!\n\n"
                f"Error: {str(e)}\n\n"
                "Please use format: 192.168.1.1-254"
            )
            self.status_label.setText(f"Error: {str(e)}")
            self.scanning = False
            self.scan_button.setText("Start Scan")
        except Exception as e:
            QMessageBox.critical(
                self, 
                "Unexpected Error", 
                f"An unexpected error occurred:\n{str(e)}"
            )
            self.status_label.setText(f"Error: {str(e)}")
            self.scanning = False
            self.scan_button.setText("Start Scan")
    
    def stop_scan(self):
        self.scanning = False
        self.scan_session_id += 1
        self.scan_button.setText("Start Scan")
        self.status_label.setText("Scan stopped")
    
    def on_scan_result(self, result):
        # Abaikan hasil dari scan session yang sudah tidak aktif
        if result.session_id != self.scan_session_id:
            return
        
        # Abaikan jika scanning sudah dihentikan
        if not self.scanning:
            return
        
        # Hanya tampilkan IP yang ONLINE
        if result.status == "Online":
            row = self.table.rowCount()
            self.table.insertRow(row)
            
            #Biar semua item di tengah
            ip_item = QTableWidgetItem(result.ip)
            ip_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 0, ip_item)
            
            hostname_item = QTableWidgetItem(result.hostname)
            hostname_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 1, hostname_item)
            
            mac_item = QTableWidgetItem(result.mac_address)
            mac_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 2, mac_item)
            
            os_type_item = QTableWidgetItem(result.os_type)
            os_type_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 3, os_type_item)
            
            status_item = QTableWidgetItem(result.status)
            status_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 4, status_item)
            
            latency_item = QTableWidgetItem(result.latency)
            latency_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 5, latency_item)
            
            self.online_count += 1
        
        #Simpan ke database
        self.db.insert_log(
            result.ip,
            result.hostname,
            result.status,
            result.latency,
            result.mac_address,
            result.os_type
        )
        
        self.scanned_count += 1
        self.progress_bar.setValue(self.scanned_count)
        
        #Update status label setiap 10 hasil
        if self.scanned_count % 10 == 0:
            self.status_label.setText(f"Scanning... {self.scanned_count}/{self.current_scan_total} ({self.online_count} online)")
        
        if self.scanned_count >= self.current_scan_total:
            self.scanning = False
            self.scan_button.setText("Start Scan")
            self.status_label.setText(f"Scan completed! Found {self.online_count} online device(s) from {self.scanned_count} scanned")
            self.load_logs()
    
    def add_to_monitor(self):
        ip = self.monitor_ip_input.text().strip()
        if not ip:
            self.monitor_status_label.setText("Error: Enter IP address")
            return
        
        #cek apakah IP sudah ada di list monitor
        if ip in self.monitor_ips:
            self.monitor_status_label.setText(f"{ip} already in monitor list")
            return
        
        self.monitor_ips.append(ip)
        
        #Nambah ke tabel monitor(dengan perataan tengah)
        row = self.monitor_table.rowCount()
        self.monitor_table.insertRow(row)
        
        ip_item = QTableWidgetItem(ip)
        ip_item.setTextAlignment(Qt.AlignCenter)
        self.monitor_table.setItem(row, 0, ip_item)
        
        for col in range(1, 6):
            item = QTableWidgetItem("-" if col != 3 else "Waiting...")
            item.setTextAlignment(Qt.AlignCenter)
            self.monitor_table.setItem(row, col, item)
        
        self.monitor_status_label.setText(f"Added {ip} to monitor list")
        self.monitor_ip_input.clear()
    #Fungsi buat ngehandle tombol Start/Stop Monitoring
    def toggle_monitoring(self):
        if not self.monitoring:
            if not self.monitor_ips:
                self.monitor_status_label.setText("Error: Add at least one IP to monitor")
                return
            
            self.monitoring = True
            self.start_monitor_button.setText("Stop Monitoring")
            interval = self.interval_input.value() * 1000  #Konversi ke milidetik(ms)
            self.monitor_timer.start(interval)
            self.monitor_status_label.setText(f"Monitoring {len(self.monitor_ips)} IP(s) every {self.interval_input.value()}s")
            
            #Jalankan scan pertama
            self.run_monitor_scan()
        else:
            self.monitoring = False
            self.monitor_timer.stop()
            self.start_monitor_button.setText("Start Monitoring")
            self.monitor_status_label.setText("Monitoring stopped")
    #Fungsi buat ngejalanin scan di monitor
    def run_monitor_scan(self):
        for ip in self.monitor_ips:
            worker = ScanWorker(ip)
            worker.signals.result.connect(self.on_monitor_result)
            self.thread_pool.start(worker)
    #Fungsi buat ngehandle hasil scan di monitor
    def on_monitor_result(self, result):
        timestamp = datetime.now().strftime("%H:%M:%S")
        
        #Fungsi Buat Update monitor table
        for row in range(self.monitor_table.rowCount()):
            if self.monitor_table.item(row, 0).text() == result.ip:
                hostname_item = QTableWidgetItem(result.hostname)
                hostname_item.setTextAlignment(Qt.AlignCenter)
                self.monitor_table.setItem(row, 1, hostname_item)
                
                mac_item = QTableWidgetItem(result.mac_address)
                mac_item.setTextAlignment(Qt.AlignCenter)
                self.monitor_table.setItem(row, 2, mac_item)
                
                status_item = QTableWidgetItem(result.status)
                status_item.setTextAlignment(Qt.AlignCenter)
                self.monitor_table.setItem(row, 3, status_item)
                
                latency_item = QTableWidgetItem(result.latency)
                latency_item.setTextAlignment(Qt.AlignCenter)
                self.monitor_table.setItem(row, 4, latency_item)
                
                time_item = QTableWidgetItem(timestamp)
                time_item.setTextAlignment(Qt.AlignCenter)
                self.monitor_table.setItem(row, 5, time_item)
                
                #Color coding berdasarkan status
                if result.status == "Online":
                    status_item.setBackground(self.monitor_table.palette().color(self.monitor_table.palette().ColorRole.Base))
                else:
                    status_item.setBackground(QColor(255, 200, 200))
                break
        
        #Simpan ke database
        self.db.insert_log(
            result.ip,
            result.hostname,
            result.status,
            result.latency,
            result.mac_address
        )
        
        #Refresh log viewer (optimasi: jangan terlalu sering refresh nanti lemot)
        if hasattr(self, '_last_log_refresh'):
            now = datetime.now()
            if (now - self._last_log_refresh).seconds < 2:
                return
        
        self.load_logs()
        self._last_log_refresh = datetime.now()
    
    def load_logs(self):
        logs = self.db.fetch_logs_with_hostname()  #Buat Nama Hostname di Log Viewer
        self.log_table.setRowCount(0)
        
        for log in logs:
            row = self.log_table.rowCount()
            self.log_table.insertRow(row)
            for col, value in enumerate(log):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignCenter)
                self.log_table.setItem(row, col, item)
    #Fungsi buat ngehandle menu konteks(klik kanan)
    def show_scanner_context_menu(self, position):
        """Show context menu for scanner table"""
        menu = QMenu()
        
        selected_rows = self.table.selectionModel().selectedRows()
        
        #Buat ngeambil item di posisi klik kanan
        item = self.table.itemAt(position)
        
        #Biar Select All selalu ada
        select_all_action = menu.addAction("✅ Select All (Ctrl+A)")
        select_all_action.triggered.connect(lambda: self.table.selectAll())
        
        menu.addSeparator()
        
        if selected_rows:
            #Aksi untuk multiple selection
            if len(selected_rows) > 1:
                add_all_action = menu.addAction(f"➕ Add All ({len(selected_rows)}) to Ping Monitor")
                add_all_action.triggered.connect(self.add_selected_to_monitor)
                
                copy_all_action = menu.addAction(f"📋 Copy All ({len(selected_rows)}) IP Addresses")
                copy_all_action.triggered.connect(self.copy_selected_scanner_ips)
            elif item:
                #Aksi untuk single selection
                row = item.row()
                ip_item = self.table.item(row, 0)
                
                if ip_item:
                    ip_address = ip_item.text()
                    
                    #Buat nambahin ke monitor
                    add_action = menu.addAction("➕ Add to Ping Monitor")
                    add_action.triggered.connect(lambda: self.add_ip_to_monitor(ip_address))
                    
                    menu.addSeparator()
                
                #Buat Nyalin ke clipboard
                copy_action = menu.addAction("📋 Copy")
                copy_action.triggered.connect(lambda: self.copy_cell_to_clipboard(item.text()))
        
        menu.exec(self.table.viewport().mapToGlobal(position))
    #Fungsi buat ngehandle menu konteks(klik kanan) di monitor table
    def show_monitor_context_menu(self, position):
        """Show context menu for monitor table"""
        menu = QMenu()
        
        selected_rows = self.monitor_table.selectionModel().selectedRows()
        
        item = self.monitor_table.itemAt(position)
        
        #Biar Select All selalu ada
        select_all_action = menu.addAction("✅ Select All (Ctrl+A)")
        select_all_action.triggered.connect(lambda: self.monitor_table.selectAll())
        
        menu.addSeparator()
        
        if selected_rows:
            #Aksi untuk multiple selection
            if len(selected_rows) > 1:
                remove_all_action = menu.addAction(f"🗑️ Remove All ({len(selected_rows)}) from Monitor")
                remove_all_action.triggered.connect(self.remove_from_monitor)
                
                copy_all_action = menu.addAction(f"📋 Copy All ({len(selected_rows)}) IP Addresses")
                copy_all_action.triggered.connect(self.copy_selected_monitor_ips)
            elif item:
                #Aksi untuk single selection - Nyalin
                copy_action = menu.addAction("📋 Copy")
                copy_action.triggered.connect(lambda: self.copy_cell_to_clipboard(item.text()))
        
        menu.exec(self.monitor_table.viewport().mapToGlobal(position))
    #Fungsi buat ngehandle menu konteks(klik kanan) di log table
    def show_log_context_menu(self, position):
        """Show context menu for log table"""
        menu = QMenu()
        
        item = self.log_table.itemAt(position)
        if item:
            #Aksi untuk Nyalin
            copy_action = menu.addAction("📋 Copy")
            copy_action.triggered.connect(lambda: self.copy_cell_to_clipboard(item.text()))
        
        menu.exec(self.log_table.viewport().mapToGlobal(position))
    #Fungsi buat nambahin IP ke monitor dari hasil scanner
    def add_ip_to_monitor(self, ip_address):
        """Add IP address to ping monitor from scanner results"""
        if (ip_address in self.monitor_ips):
            QMessageBox.information(self, "Already in Monitor", f"{ip_address} is already being monitored")
            return
        
        self.monitor_ips.append(ip_address)
        
        #Tambahkan ke tabel dengan perataan tengah
        row = self.monitor_table.rowCount()
        self.monitor_table.insertRow(row)
        
        ip_item = QTableWidgetItem(ip_address)
        ip_item.setTextAlignment(Qt.AlignCenter)
        self.monitor_table.setItem(row, 0, ip_item)
        
        for col in range(1, 6):
            item = QTableWidgetItem("-" if col != 3 else "Waiting...")
            item.setTextAlignment(Qt.AlignCenter)
            self.monitor_table.setItem(row, col, item)
        
        #Pindah ke tab monitor
        self.tab_widget.setCurrentIndex(1)
        
        QMessageBox.information(self, "Added to Monitor", f"{ip_address} has been added to Ping Monitor")
    #Fungsi buat ngehapus IP dari monitor list
    def remove_from_monitor(self):
        """Remove selected IP from monitor list"""
        selected_rows = self.monitor_table.selectionModel().selectedRows()
        
        if not selected_rows:
            QMessageBox.warning(self, "No Selection", "Please select an IP address to remove from monitor")
            return
        
        #Hapus dari list dan tabel (urutannya dibalik supaya gak error index)
        for index in sorted(selected_rows, reverse=True):
            row = index.row()
            ip_item = self.monitor_table.item(row, 0)
            if ip_item:
                ip_address = ip_item.text()
                if ip_address in self.monitor_ips:
                    self.monitor_ips.remove(ip_address)
            self.monitor_table.removeRow(row)
        
        self.monitor_status_label.setText(f"Removed IP(s) from monitor list. {len(self.monitor_ips)} IP(s) remaining")
    #Fungsi buat nyalin teks ke clipboard
    def copy_cell_to_clipboard(self, text):
        """Copy text to clipboard"""
        clipboard = QApplication.clipboard()
        clipboard.setText(text)
        
        #Nampilin pesan sukses
        if hasattr(self, 'status_label'):
            original_text = self.status_label.text()
            self.status_label.setText(f"✓ Copied: {text}")
            QTimer.singleShot(2000, lambda: self.status_label.setText(original_text))
    #Fungsi buat nambahin semua IP yang dipilih di scanner ke monitor
    def add_selected_to_monitor(self):
        """Add all selected IPs from scanner to monitor"""
        selected_rows = self.table.selectionModel().selectedRows()
        
        if not selected_rows:
            return
        
        added_count = 0
        skipped_count = 0
        
        for index in selected_rows:
            row = index.row()
            ip_item = self.table.item(row, 0)
            
            if ip_item:
                ip_address = ip_item.text()
                
                #Cek apakah sudah ada di daftar monitor
                if ip_address in self.monitor_ips:
                    skipped_count += 1
                    continue
                
                #Tambahkan ke daftar monitor
                self.monitor_ips.append(ip_address)
                
                #Tambahkan ke tabel monitor
                monitor_row = self.monitor_table.rowCount()
                self.monitor_table.insertRow(monitor_row)
                
                ip_item_copy = QTableWidgetItem(ip_address)
                ip_item_copy.setTextAlignment(Qt.AlignCenter)
                self.monitor_table.setItem(monitor_row, 0, ip_item_copy)
                
                for col in range(1, 6):
                    item = QTableWidgetItem("-" if col != 3 else "Waiting...")
                    item.setTextAlignment(Qt.AlignCenter)
                    self.monitor_table.setItem(monitor_row, col, item)
                
                added_count += 1
        
        #Tampilkan pesan ringkasan
        message = f"Added {added_count} IP(s) to Ping Monitor"
        if skipped_count > 0:
            message += f"\n{skipped_count} IP(s) were already in monitor list"
        
        self.status_label.setText(message)
        
        #Pindah ke tab monitor
        if added_count > 0:
            self.tab_widget.setCurrentIndex(1)
            QMessageBox.information(self, "Batch Add Complete", message)
    #Fungsi buat nyalin semua IP yang dipilih di scanner ke clipboard
    def copy_selected_scanner_ips(self):
        """Copy all selected IP addresses from scanner table to clipboard"""
        selected_rows = self.table.selectionModel().selectedRows()
        
        if not selected_rows:
            return
        
        ip_list = []
        for index in selected_rows:
            row = index.row()
            ip_item = self.table.item(row, 0)
            if ip_item:
                ip_list.append(ip_item.text())
        
        if ip_list:
            clipboard_text = "\n".join(ip_list)
            clipboard = QApplication.clipboard()
            clipboard.setText(clipboard_text)
            
            self.status_label.setText(f"✓ Copied {len(ip_list)} IP address(es) to clipboard")
            QTimer.singleShot(2000, lambda: self.status_label.setText("Ready"))
    #Fungsi buat nyalin semua IP yang dipilih di monitor ke clipboard
    def copy_selected_monitor_ips(self):
        """Copy all selected IP addresses from monitor table to clipboard"""
        selected_rows = self.monitor_table.selectionModel().selectedRows()
        
        if not selected_rows:
            return
        
        ip_list = []
        for index in selected_rows:
            row = index.row()
            ip_item = self.monitor_table.item(row, 0)
            if ip_item:
                ip_list.append(ip_item.text())
        
        if ip_list:
            clipboard_text = "\n".join(ip_list)
            clipboard = QApplication.clipboard()
            clipboard.setText(clipboard_text)
            
            self.monitor_status_label.setText(f"✓ Copied {len(ip_list)} IP address(es) to clipboard")
            QTimer.singleShot(2000, self._restore_monitor_status)
    #Buat nge-restore status label monitor
    def _restore_monitor_status(self):
        """Helper method to restore monitor status label"""
        if not self.monitoring:
            self.monitor_status_label.setText("Not monitoring")
        else:
            self.monitor_status_label.setText(f"Monitoring {len(self.monitor_ips)} IP(s) every {self.interval_input.value()}s")
    #Fungsi buat ngeclear semua log history
    def clear_log_history(self):
        """Clear all log history from database with confirmation"""
        #Cek apakah ada log atau gak
        log_count = self.log_table.rowCount()
        
        if log_count == 0:
            QMessageBox.information(self, "No Logs", "There are no logs to clear.")
            return
        
        #buat nampilin dialog konfirmasi
        reply = QMessageBox.warning(
            self,
            "Confirm Clear History",
            f"Are you sure you want to delete all {log_count} log entries?\n\n"
            "This action cannot be undone!",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                #Hapus semua log di database
                self.db.clear_all_logs()
                
                #Hapus semua log di tabel
                self.log_table.setRowCount(0)
                
                QMessageBox.information(
                    self, 
                    "Success", 
                    f"Successfully deleted {log_count} log entries."
                )
            except Exception as e:
                QMessageBox.critical(
                    self, 
                    "Error", 
                    f"Failed to clear log history:\n{str(e)}"
                )
    #Fungsi buat ngehandle event close aplikasi
    def closeEvent(self, event):
        """Handle application close - cleanup resources"""
        try:
            #Buat Berhentiin scan kalo lagi jalan
            if self.monitoring:
                self.monitor_timer.stop()
            
            #Tunggu semua thread selesai
            self.thread_pool.waitForDone(3000)  #Tunnggu maksimal 3 detik(3000 ms)
            
            #Berhentiin koneksi database
            self.db.close()
            
            event.accept()
        except Exception as e:
            print(f"Error during cleanup: {str(e)}")
            event.accept()
        #Aplikasi bakal nutup meskipun ada error